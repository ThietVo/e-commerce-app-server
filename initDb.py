from src import db
from src.models import Products
from sqlalchemy.sql import func
import openpyxl

db.drop_all()
db.create_all()

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("products.xlsx")
 
# Define variable to read sheet
dataframe1 = dataframe.active
countCol = 0
price = 0
# Iterate the loop to read the cell values
for row in range(0, dataframe1.max_row):
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        countCol += 1
        p = 'p' + str(countCol)
        productName = 'productName' + str(countCol)
        price += 100
        # print(col[row].value)
        p = Products(
            productName=productName,
            image=col[row].value,
            price=price,
            unit='each',
            sale='yes',
            created_at=func.now(),
            updated_at=func.now()
        )
        db.session.add_all([p])
        db.session.commit()
        